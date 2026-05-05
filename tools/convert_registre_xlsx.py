#!/usr/bin/env python3
"""Convertit le registre Excel des références en data/registre.json.

Usage :
    python tools/convert_registre_xlsx.py chemin/vers/registre.xlsx

Le script cherche les colonnes usuelles : ID, Auteur, Référence complète, Nature,
Usage, Statut, Concepts, Passages concernés, Remarques. Les noms proches sont acceptés.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ALIASES = {
    "id": ["id", "id source", "source", "n°", "numero", "numéro"],
    "auteur": ["auteur", "auteurs", "author"],
    "reference": ["référence complète", "reference complete", "référence", "reference", "titre"],
    "nature": ["nature", "type", "type de source"],
    "usage": ["usage", "usage précis", "mode d’usage", "emploi"],
    "statut": ["statut", "status", "etat", "état"],
    "concepts": ["concepts", "cadres", "cadres / concepts mobilisés"],
    "passages": ["passages concernés", "passages", "repères", "paragraphes"],
    "remarques": ["remarques", "limites", "observations"]
}


def norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace("\n", " ")


def find_header_map(headers: list[Any]) -> dict[str, int]:
    normalized = [norm(h) for h in headers]
    mapping: dict[str, int] = {}
    for target, aliases in ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                mapping[target] = normalized.index(alias)
                break
    return mapping


def cell(row: tuple[Any, ...], mapping: dict[str, int], key: str) -> str:
    idx = mapping.get(key)
    if idx is None or idx >= len(row):
        return ""
    value = row[idx]
    return "" if value is None else str(value).strip()


def convert(xlsx_path: Path, output_path: Path) -> None:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Le classeur est vide.")

    mapping = find_header_map(list(rows[0]))
    if "id" not in mapping:
        raise ValueError("Colonne ID introuvable. Le registre doit comporter une colonne 'ID' ou 'ID source'.")

    records = []
    for row in rows[1:]:
        source_id = cell(row, mapping, "id")
        if not source_id:
            continue
        records.append({
            "id": source_id,
            "auteur": cell(row, mapping, "auteur"),
            "reference": cell(row, mapping, "reference"),
            "nature": cell(row, mapping, "nature"),
            "usage": cell(row, mapping, "usage"),
            "statut": cell(row, mapping, "statut") or "à qualifier",
            "concepts": cell(row, mapping, "concepts"),
            "passages": cell(row, mapping, "passages"),
            "remarques": cell(row, mapping, "remarques")
        })

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"{len(records)} sources exportées vers {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path, help="Chemin du registre Excel")
    parser.add_argument("--output", type=Path, default=Path("data/registre.json"), help="Chemin JSON de sortie")
    args = parser.parse_args()
    convert(args.xlsx, args.output)


if __name__ == "__main__":
    main()
